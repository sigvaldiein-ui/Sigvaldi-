"""Web-layer helper utilities for Alvitur.

Sprint 71 Track A.4c — extracted from interfaces/web_server.py.
"""
import io
import logging
import os
from fastapi import HTTPException

logger = logging.getLogger("alvitur.web")

async def _polish_fn_txt(*args, **kwargs) -> str:
    """Async no-op polish stub — Sprint 63 Track A6.2.
    Caller notar 'await' svo þetta er async. Skilar textanum óbreyttum.
    TODO: bæta við raunverulegri íslenskri málfræðipúlisingu síðar.
    """
    text = args[0] if args else kwargs.get("text", "")
    return text if isinstance(text, str) else str(text)

def _detect_filetype(data: bytes, filename: str) -> str:
    """Return 'pdf', 'docx', 'xlsx', or raise HTTPException."""
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    # CSV og plain text skrár hafa engan magic byte — extension-only check
    if ext == 'csv':
        return 'csv'
    if ext == 'xls':
        return 'xls'
    header = data[:4]
    if header == b'%PDF':
        if ext != 'pdf':
            raise HTTPException(status_code=415,
                detail="Skráin er merkt sem PDF en innihald stemmir ekki.")
        return 'pdf'
    if header[:2] == b'PK':
        if ext == 'docx':
            return 'docx'
        if ext == 'xlsx':
            return 'xlsx'
        if ext == 'csv':
            return 'csv'
        if ext == 'xls':
            return 'xls'
        raise HTTPException(status_code=415,
            detail="Office skjal þekkt en skráarending er óþêkkt. Sendu .docx eða .xlsx.")
    raise HTTPException(status_code=415,
        detail="Skráargerð ekki stuðd. Styður PDF, Word (.docx), Excel (.xlsx) og CSV.")


def _parse_docx(data: bytes) -> tuple[int, list[str]]:
    """Extract text from .docx. Returns (page_estimate, text_parts)."""
    from docx import Document
    import io
    doc = Document(io.BytesIO(data))
    parts = []
    for i, para in enumerate(doc.paragraphs):
        t = para.text.strip()
        if t:
            parts.append(t)
    # Estimate pages: ~3000 chars per page
    total_chars = sum(len(p) for p in parts)
    page_estimate = max(1, total_chars // 3000)
    return page_estimate, parts


def _parse_xlsx(data: bytes) -> tuple[int, list[str]]:
    """Extract text from .xlsx. Returns (sheet_count, text_parts)."""
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    sheet_count = len(wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows_text.append(" | ".join(cells))
                row_count += 1
            if row_count >= 500:  # cap at 500 rows per sheet
                rows_text.append("[... fleiri raðir ...]")
                break
        if rows_text:
            parts.append(f"[Blað: {sheet_name}]\n" + "\n".join(rows_text))
    wb.close()
    return sheet_count, parts

# ─────────────────────────────────────────────────────────────────────────────
# ─ Sprint 27: S4 Wallet Circuit Breaker ───────────────────────────────────


def _estimate_tokens(text):
    return int(len((text or "").split()) * 1.3)

