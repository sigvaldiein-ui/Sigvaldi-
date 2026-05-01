"""
Sprint 62 Patch B: Use pandas preprocess_excel() for xlsx uploads.

Old behavior: flat pipe-separated text → LLM hallucinates sums.
New behavior: pandas computes real sums → LLM just explains.
Fallback to openpyxl flat text if pandas fails.
"""
import sys
from pathlib import Path

TARGET = Path("interfaces/web_server.py")

OLD = '''            if _filetype == 'xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(efni), read_only=True, data_only=True)
                txt = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        clean = [str(c) if c is not None else "" for c in row]
                        line = " | ".join(v for v in clean if v.strip())
                        if line: txt.append(line)
                heildartexti = "\\n".join(txt)
                sidur = len(wb.worksheets)'''

NEW = '''            if _filetype == 'xlsx':
                # Sprint 62 Patch B: Pandas pre-processor (Reiknivélar-Agent).
                # Returns markdown with real computed sums so LLM doesn't hallucinate.
                try:
                    from interfaces.excel_preprocessor import preprocess_excel as _prep_xlsx
                    heildartexti = _prep_xlsx(efni)
                    sidur = 1
                    logger.info(f"[ALVITUR] Sprint62b xlsx preprocessed via pandas, {len(heildartexti)} chars")
                except Exception as _xe:
                    logger.warning(f"[ALVITUR] pandas preprocess failed, fallback to openpyxl flat: {type(_xe).__name__}: {_xe}")
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(efni), read_only=True, data_only=True)
                    txt = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            clean = [str(c) if c is not None else "" for c in row]
                            line = " | ".join(v for v in clean if v.strip())
                            if line: txt.append(line)
                    heildartexti = "\\n".join(txt)
                    sidur = len(wb.worksheets)'''

def main():
    src = TARGET.read_text()
    if "Sprint 62 Patch B" in src:
        print("⚠️  Patch B þegar beitt.")
        return 0
    if OLD not in src:
        print("❌ Fann ekki gamla strenginn.")
        return 1
    if src.count(OLD) > 1:
        print(f"❌ Strengur fannst {src.count(OLD)} sinnum.")
        return 1
    patched = src.replace(OLD, NEW)
    import ast
    try:
        ast.parse(patched)
    except SyntaxError as e:
        print(f"❌ Syntax villa: {e}")
        return 1
    TARGET.write_text(patched)
    print("✅ Patch B beittur. Syntax OK.")
    return 0

if __name__ == "__main__":
    sys.exit(main())
